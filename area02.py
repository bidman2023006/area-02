import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook

try:

    wb = load_workbook('清單.xlsx', data_only=True)

    sheet = wb.active
    print(sheet.max_row)
    print(sheet.max_column)

    A_LIST = []
    E_LIST = []
    G_LIST = []


    def delete_extra_zero(n):
        """删除小数点后多余的0"""
        n = '{:g}'.format(n)
        n = float(n) if '.' in n else int(n)  # 含小数点转float否则int
        return n


    for i in range(30):
        try:
            A = sheet.cell(row=2, column=8 + i).value
            A_LIST.append(delete_extra_zero(int(A) / 10))
            E = sheet.cell(row=3, column=8 + i).value
            E_LIST.append(delete_extra_zero(int(E) / 10))
            G = sheet.cell(row=4, column=8 + i).value
            G_LIST.append(delete_extra_zero(int(G) / 10))
        except:
            pass

    print(A_LIST)
    print(E_LIST)

    SKU_DICT = {}
    for i in range(sheet.max_row):
        sku = sheet.cell(row=1 + i, column=1).value
        # print(sku)
        if len(sku) == 9:
            OK_sku = '{}-{}'.format(sku[:-3], sku[-3:])
            price = sheet.cell(row=1 + i, column=3).value
            # SKU_DICT[OK_sku]={'price':int(price)}

            NB_LIST = []
            for AA in range(len(A_LIST)):
                NB = sheet.cell(row=1 + i, column=8 + AA).value
                if NB == '' or NB == None:
                    NB = 0
                if sheet.cell(row=1 + i, column=7).value == 'A-A':
                    NB_LIST.append('{}:{}'.format(A_LIST[AA], NB))
                if sheet.cell(row=1 + i, column=7).value == 'E-E':
                    NB_LIST.append('{}:{}'.format(E_LIST[AA], NB))
                if sheet.cell(row=1 + i, column=7).value == 'G-G':
                    NB_LIST.append('{}:{}'.format(G_LIST[AA], NB))
            SKU_DICT[OK_sku] = {'NB': NB_LIST, 'price': int(round(price))}

    print(SKU_DICT)
    path = 'TEST.txt'
    f = open(path, 'w')
    for key in SKU_DICT:
        sku = key
        price = SKU_DICT[key]['price']
        for i in range(len(SKU_DICT[key]['NB'])):
            size = SKU_DICT[key]['NB'][i].split(':')[0]
            NNBB = SKU_DICT[key]['NB'][i].split(':')[1]
            OKKK = '{}\t{}\t{}\t{}'.format(sku, price, size, NNBB)
            print(OKKK)
            f.write(OKKK + '\n')

    input('\nSUCCESS')
except:
    input('\nERROR')
